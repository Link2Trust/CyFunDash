#!/usr/bin/env python3
"""
Parse the CyFun2025 Self-Assessment Excel file and extract structured data as JSON.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

EXCEL_FILE = Path(__file__).parent / "CyFun2025_Self-Assessment_tool_ESSENTIAL_v3.1.xlsx"
OUTPUT_FILE = Path(__file__).parent / "data.json"

FUNCTION_SHEETS = ["GOVERN", "IDENTIFY", "PROTECT", "DETECT", "RESPOND", "RECOVER"]

ASSURANCE_HIERARCHY = {"Basic": 1, "Important": 2, "Essential": 3}


def _get_cell(row, index):
    """Get cell value at index, or None if out of range."""
    return row[index] if len(row) > index else None


def _parse_subcategory_text(text):
    """Split 'CODE: Description' into (code, description)."""
    if ":" in text:
        code, desc = text.split(":", 1)
        return code.strip(), desc.strip()
    return text, ""


def _build_requirement(col_c, col_e, col_f):
    """Build a requirement dict from column values."""
    is_key_measure = bool(col_c and str(col_c).strip().lower() == "key measure")
    return {
        "assurance_level": col_e.strip() if isinstance(col_e, str) else str(col_e),
        "requirement": col_f.strip(),
        "key_measure": is_key_measure,
    }


def parse_function_sheet(ws, sheet_name):
    """Parse a function sheet (GOVERN, IDENTIFY, etc.) and return structured data."""
    categories = []
    current_category = None
    current_subcategory = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        col_a = _get_cell(row, 0)
        col_c = _get_cell(row, 2)
        col_d = _get_cell(row, 3)
        col_e = _get_cell(row, 4)
        col_f = _get_cell(row, 5)

        # New category
        if col_a and isinstance(col_a, str) and col_a.strip():
            current_category = {"name": col_a.strip(), "subcategories": []}
            categories.append(current_category)

        # New subcategory
        if col_d and isinstance(col_d, str) and col_d.strip():
            code, desc = _parse_subcategory_text(col_d.strip())
            current_subcategory = {"code": code, "description": desc, "requirements": []}
            if current_category:
                current_category["subcategories"].append(current_subcategory)

        # Requirement row
        if col_e and col_f and isinstance(col_f, str) and col_f.strip() and current_subcategory:
            current_subcategory["requirements"].append(
                _build_requirement(col_c, col_e, col_f)
            )

    return categories


def parse_maturity_levels(wb):
    """Parse the Maturity Levels sheet."""
    ws = wb["Maturity Levels"]
    levels = []
    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        if row[0] and row[1] is not None:
            levels.append({
                "name": str(row[0]).strip(),
                "value": row[1],
                "documentation": str(row[2]).strip() if row[2] else "",
                "implementation": str(row[3]).strip() if row[3] else "",
            })
    return levels


# Column groups for key measures: (code_col, desc_col, target_col)
_KM_COLUMN_GROUPS = [
    (11, 12, 13),  # L, M, N
    (18, 19, 20),  # S, T, U
    (25, 26, 27),  # Z, AA, AB
]


def _parse_summary_categories(ws):
    """Parse category entries from the ESSENTIAL Summary sheet."""
    categories = []
    current_function = None
    for row in ws.iter_rows(min_row=4, max_row=24, values_only=True):
        col_a = _get_cell(row, 0)
        col_b = _get_cell(row, 1)
        col_c = _get_cell(row, 2)
        if col_a and isinstance(col_a, str):
            current_function = col_a.strip()
        if col_b and isinstance(col_b, str) and col_b.strip():
            categories.append({
                "function": current_function,
                "category": col_b.strip(),
                "target_maturity": col_c if col_c else 0,
            })
    return categories


def _extract_km(row, code_col, desc_col, target_col, pattern):
    """Try to extract a key measure from specific columns in a row."""
    if len(row) <= desc_col:
        return None
    cell = row[code_col]
    if not (cell and isinstance(cell, str)):
        return None
    code = str(cell).strip()
    if not pattern.match(code):
        return None
    target = row[target_col] if len(row) > target_col and row[target_col] else 3
    return {
        "code": code,
        "description": str(row[desc_col]).strip() if row[desc_col] else "",
        "target_maturity": target,
    }


def _parse_key_measures(ws, pattern):
    """Parse key measures from all column groups in the summary sheet."""
    key_measures = []
    for row in ws.iter_rows(min_row=26, max_row=ws.max_row, values_only=True):
        for code_col, desc_col, target_col in _KM_COLUMN_GROUPS:
            km = _extract_km(row, code_col, desc_col, target_col, pattern)
            if km:
                key_measures.append(km)
    return key_measures


def parse_summary(wb):
    """Parse the ESSENTIAL Summary sheet for category scores and key measures."""
    ws = wb["ESSENTIAL Summary"]

    target_total = None
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        target_total = _get_cell(row, 7)

    categories = _parse_summary_categories(ws)
    km_code_pattern = re.compile(r'^[A-Z]{2}\.[A-Z]{2}-\d')
    key_measures = _parse_key_measures(ws, km_code_pattern)

    return {
        "categories": categories,
        "target_total_maturity": target_total if target_total else 3.5,
        "key_measures": key_measures,
    }


def _compute_statistics(functions):
    """Compute per-function requirement statistics."""
    stats = {}
    for fn_name, categories in functions.items():
        fn_stats = {"Basic": 0, "Important": 0, "Essential": 0, "total": 0}
        for cat in categories:
            for sub in cat["subcategories"]:
                for req in sub["requirements"]:
                    level = req["assurance_level"]
                    if level in fn_stats:
                        fn_stats[level] += 1
                    fn_stats["total"] += 1
        stats[fn_name] = fn_stats
    return stats


def main():
    print(f"Reading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)

    data = {
        "title": "CyFun® 2025 CyberFundamentals",
        "assurance_levels": [
            {"name": "Basic", "value": 1, "description": "Basic cybersecurity hygiene"},
            {"name": "Important", "value": 2, "description": "Standard cybersecurity controls"},
            {"name": "Essential", "value": 3, "description": "Advanced cybersecurity requirements"},
        ],
        "maturity_levels": parse_maturity_levels(wb),
        "functions": {},
    }

    summary_data = parse_summary(wb)
    data["summary"] = summary_data["categories"]
    data["target_total_maturity"] = summary_data["target_total_maturity"]
    data["key_measures"] = summary_data["key_measures"]

    for sheet_name in FUNCTION_SHEETS:
        print(f"  Parsing {sheet_name}...")
        ws = wb[sheet_name]
        data["functions"][sheet_name] = parse_function_sheet(ws, sheet_name)

    stats = _compute_statistics(data["functions"])
    data["statistics"] = stats

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\nData written to {OUTPUT_FILE}")
    total = sum(s["total"] for s in stats.values())
    print(f"Total requirements extracted: {total}")
    for fn, s in stats.items():
        print(f"  {fn}: {s['total']} (Basic: {s['Basic']}, Important: {s['Important']}, Essential: {s['Essential']})")
    print(f"Key measures extracted: {len(data['key_measures'])}")
    print(f"Target total maturity: {data['target_total_maturity']}")


if __name__ == "__main__":
    main()
